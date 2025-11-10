#!/usr/bin/env python3
"""
Create Shared Document Catalog Excel Template

This script generates the Excel workbook following the specification in
SHARED_DOCUMENT_CATALOG_EXCEL_TEMPLATE.md

Requirements:
    pip install openpyxl

Usage:
    python create_excel_template.py [output_file.xlsx]
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle


def create_excel_template(output_file: str = "Shared_Document_Eligibility_Catalog.xlsx"):
    """Create the Excel template workbook."""

    print("Creating Shared Document Eligibility Catalog Excel Template...")
    print(f"Output file: {output_file}\n")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create all sheets
    print("Creating sheets...")
    create_instructions_sheet(wb)
    create_document_list_sheet(wb)
    create_document_details_sheet(wb)
    create_eligibility_conditions_sheet(wb)
    create_data_sources_sheet(wb)
    create_test_scenarios_sheet(wb)
    create_lookups_sheet(wb)

    # Set active sheet to INSTRUCTIONS
    wb.active = wb['INSTRUCTIONS']

    # Save workbook
    print(f"\nSaving workbook to: {output_file}")
    wb.save(output_file)
    print("[OK] Excel template created successfully!")
    print(f"\nNext steps:")
    print(f"  1. Open {output_file}")
    print(f"  2. Review INSTRUCTIONS sheet")
    print(f"  3. Start filling in DOCUMENT_LIST")
    print(f"  4. Use convert_excel_to_yaml.py to export")


def create_instructions_sheet(wb):
    """Create INSTRUCTIONS sheet."""
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.sheet_properties.tabColor = "4472C4"  # Blue

    # Title
    ws['A1'] = "SHARED DOCUMENT ELIGIBILITY CATALOG"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = "How to Use This Workbook"
    ws['A2'].font = Font(size=14, bold=True)

    instructions = [
        ("", ""),
        ("STEP 1: Add all shared documents to DOCUMENT_LIST sheet", ""),
        ("  - List each shared document type (Privacy Policy, Cardholder Agreement, etc.)", ""),
        ("  - Assign a unique Document ID (e.g., DOC-001, DOC-002)", ""),
        ("  - Mark priority (High/Medium/Low)", ""),
        ("", ""),
        ("STEP 2: Complete DOCUMENT_DETAILS for each document", ""),
        ("  - Fill in metadata, business context, ownership", ""),
        ("  - Select sharing scope from dropdown", ""),
        ("  - Use dropdowns wherever possible", ""),
        ("", ""),
        ("STEP 3: Define eligibility conditions in ELIGIBILITY_CONDITIONS", ""),
        ("  - For custom rules only", ""),
        ("  - One row per condition", ""),
        ("  - Link to Document ID", ""),
        ("  - Select operator from dropdown", ""),
        ("", ""),
        ("STEP 4: Map data sources in DATA_SOURCES", ""),
        ("  - Identify which APIs are needed", ""),
        ("  - Specify endpoint and response fields", ""),
        ("  - Note cache requirements", ""),
        ("", ""),
        ("STEP 5: Create test scenarios in TEST_SCENARIOS", ""),
        ("  - At least 2 scenarios per document (positive & negative)", ""),
        ("  - Include edge cases", ""),
        ("  - Document expected results", ""),
        ("", ""),
        ("REFERENCE:", ""),
        ("  - See LOOKUPS sheet for dropdown values", ""),
        ("  - Colored cells indicate required fields", ""),
        ("  - Use data validation to prevent errors", ""),
        ("  - Export to YAML/JSON for implementation", ""),
        ("", ""),
        ("SUPPORT:", ""),
        ("  Questions? Contact: technical-team@documenthub.com", ""),
        ("  Documentation: docs/guides/SHARED_DOCUMENT_ELIGIBILITY_CATALOG.md", ""),
    ]

    for idx, (text, _) in enumerate(instructions, start=3):
        ws[f'A{idx}'] = text
        if text.startswith("STEP") or text.startswith("REFERENCE:") or text.startswith("SUPPORT:"):
            ws[f'A{idx}'].font = Font(bold=True, size=12)

    # Set column widths
    ws.column_dimensions['A'].width = 80

    print("  [OK] INSTRUCTIONS sheet created")


def create_document_list_sheet(wb):
    """Create DOCUMENT_LIST sheet."""
    ws = wb.create_sheet("DOCUMENT_LIST")
    ws.sheet_properties.tabColor = "70AD47"  # Green

    # Headers
    headers = [
        "Document ID", "Document Name", "Document Type", "Line of Business",
        "Sharing Scope", "Regulatory?", "Owner Department", "Status",
        "Priority", "Estimated Volume (%)", "SME Name", "Interview Date", "Notes"
    ]

    add_header_row(ws, headers)

    # Set column widths
    widths = [15, 40, 20, 25, 25, 12, 20, 18, 12, 20, 25, 15, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Add data validation (will be added after LOOKUPS sheet is created)

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Add sample row
    sample = ["DOC-001", "Example Document", "", "", "", "", "", "", "", "", "", "", ""]
    for idx, value in enumerate(sample, start=1):
        ws.cell(row=2, column=idx, value=value)

    print("  [OK] DOCUMENT_LIST sheet created")


def create_document_details_sheet(wb):
    """Create DOCUMENT_DETAILS sheet."""
    ws = wb.create_sheet("DOCUMENT_DETAILS")
    ws.sheet_properties.tabColor = "FFC000"  # Yellow

    headers = [
        "Document ID", "Document Name", "Category", "Subcategory", "Description",
        "Regulatory", "Regulation Name", "Effective Date", "Valid Until", "Owner Dept",
        "Contact Person", "Business Justification", "Target Audience", "Expected Volume",
        "Expected %", "Revenue Opportunity", "Compliance Requirement", "Audit Logging Required",
        "Version", "Created By", "Created Date", "Approved By", "Approved Date",
        "Last Modified", "Implementation Status", "Notes"
    ]

    add_header_row(ws, headers)

    # Set column widths
    widths = [15, 40, 20, 20, 60, 12, 30, 15, 15, 20, 25, 80, 40, 15, 10, 20, 40, 15,
              10, 25, 15, 25, 15, 15, 20, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else f'A{chr(64 + idx - 26)}'].width = width

    # Add formulas for row 2
    ws['B2'] = '=IFERROR(VLOOKUP(A2,DOCUMENT_LIST!$A:$B,2,FALSE),"")'
    ws['O2'] = '=IF(N2="","",N2)'
    ws['X2'] = f'=TODAY()'

    ws.freeze_panes = 'A2'

    print("  [OK] DOCUMENT_DETAILS sheet created")


def create_eligibility_conditions_sheet(wb):
    """Create ELIGIBILITY_CONDITIONS sheet."""
    ws = wb.create_sheet("ELIGIBILITY_CONDITIONS")
    ws.sheet_properties.tabColor = "ED7D31"  # Orange

    headers = [
        "Condition ID", "Document ID", "Document Name", "Condition Priority",
        "Condition Description", "Field Name", "Data Source", "Operator",
        "Value", "Value Type", "Logical Operator", "Required?",
        "Error Message", "Example (Pass)", "Example (Fail)", "Notes"
    ]

    add_header_row(ws, headers)

    widths = [15, 15, 40, 10, 60, 25, 25, 25, 30, 15, 15, 12, 50, 40, 40, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Add formula for Document Name
    ws['C2'] = '=IFERROR(VLOOKUP(B2,DOCUMENT_LIST!$A:$B,2,FALSE),"")'

    ws.freeze_panes = 'A2'

    print("  [OK] ELIGIBILITY_CONDITIONS sheet created")


def create_data_sources_sheet(wb):
    """Create DATA_SOURCES sheet."""
    ws = wb.create_sheet("DATA_SOURCES")
    ws.sheet_properties.tabColor = "7030A0"  # Purple

    headers = [
        "Data Source ID", "Data Source Name", "Source Type", "API Endpoint",
        "HTTP Method", "Request Body", "Timeout (ms)", "Retry Attempts",
        "Field Name", "JSON Path", "Data Type", "Cache Enabled",
        "Cache TTL (sec)", "Response Time (ms)", "Owner Team", "API Documentation", "Notes"
    ]

    add_header_row(ws, headers)

    widths = [20, 30, 20, 80, 15, 60, 15, 15, 25, 40, 15, 15, 15, 15, 25, 60, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws.freeze_panes = 'A2'

    print("  [OK] DATA_SOURCES sheet created")


def create_test_scenarios_sheet(wb):
    """Create TEST_SCENARIOS sheet."""
    ws = wb.create_sheet("TEST_SCENARIOS")
    ws.sheet_properties.tabColor = "C00000"  # Red

    headers = [
        "Scenario ID", "Document ID", "Document Name", "Scenario Type",
        "Scenario Name", "Test Data (JSON)", "Expected Result", "Reason",
        "Validation Steps", "Tested?", "Test Date", "Test Result",
        "Tester Name", "Notes"
    ]

    add_header_row(ws, headers)

    widths = [15, 15, 40, 20, 50, 100, 20, 80, 100, 12, 15, 15, 25, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Add formula for Document Name
    ws['C2'] = '=IFERROR(VLOOKUP(B2,DOCUMENT_LIST!$A:$B,2,FALSE),"")'

    ws.freeze_panes = 'A2'

    print("  [OK] TEST_SCENARIOS sheet created")


def create_lookups_sheet(wb):
    """Create LOOKUPS sheet with all reference data."""
    ws = wb.create_sheet("LOOKUPS")
    ws.sheet_properties.tabColor = "808080"  # Gray

    lookups = {
        'A': ('DOCUMENT_TYPE', [
            'disclosure', 'agreement', 'notice', 'statement', 'policy',
            'terms_and_conditions', 'regulatory', 'marketing', 'benefits', 'fees'
        ]),
        'B': ('LINE_OF_BUSINESS', [
            'credit_card', 'digital_banking', 'enterprise', 'all',
            'savings', 'checking', 'loan', 'investment'
        ]),
        'C': ('SHARING_SCOPE', [
            'all', 'credit_card_account_only', 'digital_bank_customer_only',
            'enterprise_customer_only', 'custom_rule'
        ]),
        'D': ('YES_NO', ['Yes', 'No']),
        'E': ('OWNER_DEPT', [
            'Legal', 'Compliance', 'Marketing', 'Product', 'Operations',
            'IT', 'Risk Management', 'Customer Service'
        ]),
        'F': ('STATUS', [
            'Draft', 'In Review', 'Approved', 'Rejected',
            'Implemented', 'Live', 'Deprecated'
        ]),
        'G': ('PRIORITY', ['High', 'Medium', 'Low']),
        'H': ('OPERATOR', [
            'EQUALS', 'NOT_EQUALS', 'GREATER_THAN', 'LESS_THAN',
            'GREATER_THAN_OR_EQUALS', 'LESS_THAN_OR_EQUALS',
            'IN', 'NOT_IN', 'CONTAINS', 'NOT_CONTAINS',
            'STARTS_WITH', 'ENDS_WITH', 'IS_NULL', 'IS_NOT_NULL',
            'DATE_BEFORE', 'DATE_AFTER', 'MONTH_EQUALS', 'REGEX_MATCH'
        ]),
        'I': ('VALUE_TYPE', ['String', 'Number', 'Boolean', 'Date', 'Array', 'Object']),
        'J': ('SCENARIO_TYPE', ['Positive', 'Negative', 'Edge Case', 'Error Handling']),
        'K': ('SOURCE_TYPE', ['REST_API', 'DATABASE', 'CACHE']),
        'L': ('HTTP_METHOD', ['GET', 'POST', 'PUT', 'DELETE']),
        'M': ('LOGICAL_OPERATOR', ['AND', 'OR']),
        'N': ('IMPL_STATUS', ['Not Started', 'In Progress', 'Testing', 'Live', 'Deferred'])
    }

    # Add headers and data
    for col, (header, values) in lookups.items():
        ws[f'{col}1'] = header
        ws[f'{col}1'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}1'].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        for idx, value in enumerate(values, start=2):
            ws[f'{col}{idx}'] = value

        ws.column_dimensions[col].width = 30

    ws.freeze_panes = 'A2'

    # Now add data validation to other sheets
    add_data_validations(wb)

    print("  [OK] LOOKUPS sheet created")


def add_data_validations(wb):
    """Add data validation dropdowns to all sheets."""

    # DOCUMENT_LIST validations
    ws_list = wb['DOCUMENT_LIST']

    # Document Type (Column C)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$A$2:$A$20", allow_blank=False)
    dv.add(f'C2:C1000')
    ws_list.add_data_validation(dv)

    # Line of Business (Column D)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$B$2:$B$20", allow_blank=False)
    dv.add(f'D2:D1000')
    ws_list.add_data_validation(dv)

    # Sharing Scope (Column E)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$C$2:$C$10", allow_blank=False)
    dv.add(f'E2:E1000')
    ws_list.add_data_validation(dv)

    # Regulatory (Column F)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$D$2:$D$3", allow_blank=False)
    dv.add(f'F2:F1000')
    ws_list.add_data_validation(dv)

    # Owner Department (Column G)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$E$2:$E$20", allow_blank=False)
    dv.add(f'G2:G1000')
    ws_list.add_data_validation(dv)

    # Status (Column H)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$F$2:$F$10", allow_blank=False)
    dv.add(f'H2:H1000')
    ws_list.add_data_validation(dv)

    # Priority (Column I)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$G$2:$G$4", allow_blank=False)
    dv.add(f'I2:I1000')
    ws_list.add_data_validation(dv)

    # ELIGIBILITY_CONDITIONS validations
    ws_cond = wb['ELIGIBILITY_CONDITIONS']

    # Operator (Column H)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$H$2:$H$25", allow_blank=False)
    dv.add(f'H2:H1000')
    ws_cond.add_data_validation(dv)

    # Value Type (Column J)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$I$2:$I$10", allow_blank=False)
    dv.add(f'J2:J1000')
    ws_cond.add_data_validation(dv)

    # Logical Operator (Column K)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$M$2:$M$3", allow_blank=False)
    dv.add(f'K2:K1000')
    ws_cond.add_data_validation(dv)

    # Required (Column L)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$D$2:$D$3", allow_blank=False)
    dv.add(f'L2:L1000')
    ws_cond.add_data_validation(dv)

    # DATA_SOURCES validations
    ws_ds = wb['DATA_SOURCES']

    # Source Type (Column C)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$K$2:$K$4", allow_blank=False)
    dv.add(f'C2:C1000')
    ws_ds.add_data_validation(dv)

    # HTTP Method (Column E)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$L$2:$L$5", allow_blank=False)
    dv.add(f'E2:E1000')
    ws_ds.add_data_validation(dv)

    # Data Type (Column K)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$I$2:$I$10", allow_blank=False)
    dv.add(f'K2:K1000')
    ws_ds.add_data_validation(dv)

    # Cache Enabled (Column L)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$D$2:$D$3", allow_blank=False)
    dv.add(f'L2:L1000')
    ws_ds.add_data_validation(dv)

    # TEST_SCENARIOS validations
    ws_test = wb['TEST_SCENARIOS']

    # Scenario Type (Column D)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$J$2:$J$5", allow_blank=False)
    dv.add(f'D2:D1000')
    ws_test.add_data_validation(dv)

    # Expected Result (Column G)
    dv = DataValidation(type="list", formula1='"SHOW,DO NOT SHOW"', allow_blank=False)
    dv.add(f'G2:G1000')
    ws_test.add_data_validation(dv)

    # Tested (Column J)
    dv = DataValidation(type="list", formula1="=LOOKUPS!$D$2:$D$3", allow_blank=True)
    dv.add(f'J2:J1000')
    ws_test.add_data_validation(dv)

    # Test Result (Column L)
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked"', allow_blank=True)
    dv.add(f'L2:L1000')
    ws_test.add_data_validation(dv)

    print("  [OK] Data validations added")


def add_header_row(ws, headers):
    """Add formatted header row."""
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add border
        thin_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        cell.border = thin_border

    ws.row_dimensions[1].height = 30


def main():
    """Main entry point."""
    if len(sys.argv) > 1:
        output_file = sys.argv[1]
    else:
        output_file = "Shared_Document_Eligibility_Catalog.xlsx"

    # Check if file exists
    if Path(output_file).exists():
        response = input(f"\n{output_file} already exists. Overwrite? (y/n): ")
        if response.lower() != 'y':
            print("Cancelled.")
            sys.exit(0)

    print("=" * 80)
    print("  Shared Document Catalog - Excel Template Generator")
    print("=" * 80)
    print()

    try:
        create_excel_template(output_file)

        print()
        print("=" * 80)
        print("  Excel Template Created Successfully!")
        print("=" * 80)
        print()

    except Exception as e:
        print()
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
